# Naceteni dat z txt soboru a rozdeleni na jednotlive linky kusovniku. Vysledek ulozen jako list.
from operator import contains
import openpyxl as excel
import os


def nacteni_dat(file):
    with open(file) as datafile:
        data = datafile.readlines()
        datafile.close()
    vse_ocistene = []
    for linka in data:
        if linka[0][0] == "|":
            slinka = linka.split("|")
            linka_bez_mezer = []
            for data in slinka:
                data = data.strip()
                data = data.replace(",", "")
                if data == "":
                    data = data.replace("", "0")
                linka_bez_mezer.append(data)
            vse_ocistene.append(linka_bez_mezer)
    return vse_ocistene

# Vytvoreni kusovniku itemu jako list po jednotlivych linkach kusovniku.
def vytvoreni_kusovniku(data):
    kusovnik_vseho = []  # Kusovnik vsech linek:
    for line in data:
        if line[17] == "19-JAN-38":
            i = 0
            kusovnik_linky = []
            item = line[1]
            while item != "0":
                kusovnik_linky.append(item)
                i += 12
                item = line[1+i]
            if kusovnik_linky not in kusovnik_vseho:
                kusovnik_vseho.append(kusovnik_linky)
    return kusovnik_vseho


# Pro kazdou linku vezme itemy z linky a ulozi jejich hodnoty ex date, lt, ss a routing LT jako dict do listu all parameters.
def databaze_parametru(data):
    all_parameters = {}
    uz_projeto = []
    for line in data:
        i = 0
        item = line[1]
        while item != "0":          
            item_parameters = {}
            # Pokud item jeste neresen NEBO ( item uz resen, ale nemel platny exdate                                                                        ) NEBO (item uz resen, ale nemel zadny Warehouse                                         A   ted uz naky ma            )            
            if (item not in uz_projeto) or ((item in uz_projeto) and (all_parameters.get(item).get("exdate") != "19-JAN-38") and (line[5 + i] == "19-JAN-38")) or ((item in uz_projeto) and (all_parameters.get(item).get("warehouse") == "0") and (line[4 + i] != "0")):
                item_parameters["description"] = line[2 + i]               
                item_parameters["typ"] = line[3 + i]
                item_parameters["warehouse"] = line[4 + i]
                item_parameters["exdate"] = line[5 + i]
                item_parameters["supplytime"] = float(line[6 + i])
                item_parameters["supplier"] = line[7 + i]
                item_parameters["nakupci"] = line[8 + i]
                item_parameters["safetystock"] = float(line[9 + i])
                if (item[0:3] == "PMP") and (line[10 + i] != "0"): # Priradi projektovy LT k anonymni polozce, pokud uz tam neni a neni projektovy LT 0.
                    if all_parameters.get(item[9:len(item)]) == None: # Pokud jeste neni anonym v parametrech.
                        all_parameters[item[9:len(item)]] = {"routinglt" : float(line[10 + i])}
                        item_parameters["routinglt"] = float(line[10 + i])
                    else: # Pokud uz je anonym v parametrech, prepise.
                        all_parameters[item[9:len(item)]]["routinglt"] = float(line[10 + i])
                        item_parameters["routinglt"] = float(line[10 + i])
                elif all_parameters.get(item) != None: # Pokud uz anonymni dil ma v parametrech LT z projektoveho.
                    if (line[9 + i] == "0"): # Neprepise se jestli je routing 0.
                        item_parameters["routinglt"] = all_parameters.get(item).get("routinglt")
                    elif (line[9 + i] != "0"): # Prepise se jestli neni routing 0.
                        item_parameters["routinglt"] = float(line[10 + i])
                elif all_parameters.get(item) == None: # Pro ostatni pripady vezme LT z import dat.
                    item_parameters["routinglt"] = float(line[10 + i])
                if (item[0:3] == "PMP") and (line[10 + i] != "0"): # Priradi projektovy std LT k anonymni polozce, pokud uz tam neni a neni projektovy LT 0.
                    if all_parameters.get(item[9:len(item)]) == None: # Pokud jeste neni anonym v parametrech.
                        all_parameters[item[9:len(item)]] = {"standardroutinglt" : float(line[11 + i])}
                        item_parameters["standardroutinglt"] = float(line[11 + i])
                    else: # Pokud uz je anonym v parametrech, prepise.
                        all_parameters[item[9:len(item)]]["standardroutinglt"] = float(line[11 + i])
                        item_parameters["standardroutinglt"] = float(line[11 + i])
                elif all_parameters.get(item) != None: # Pokud uz anonymni dil ma v parametrech LT z projektoveho.
                    if (line[10 + i] == "0"): # Neprepise se jestli je STDrouting 0.
                        item_parameters["standardroutinglt"] = all_parameters.get(item).get("standardroutinglt")
                    elif (line[10 + i] != "0"): # Prepise se jestli neni STDrouting 0.
                        item_parameters["standardroutinglt"] = float(line[11 + i])
                elif all_parameters.get(item) == None: # Pro ostatni pripady vezme LT z import dat.
                    item_parameters["standardroutinglt"] = float(line[11 + i])
                item_parameters["phantom"] = (line[12 + i])
                all_parameters[item] = item_parameters
                uz_projeto.append(item)
                i += 12
                item = line[1 + i]
            else:
                i += 12
                item = line[1 + i]
    return all_parameters


def item_typ(item, parameters): 
    if (item[0:3] == "314") and (parameters.get(item).get("typ") == "Manufactured"):
        typ = "panel"    
    elif item[0:3] == "PMP":
        typ = "M"
    elif (parameters.get(item).get("supplier") == "I00000008") and ((item[0] != "3") and (item[0] != "9" )): # Jedna se o nakupovany dil z Lamphunu a neni to surovy material.
        if parameters.get(item).get("warehouse") != "PZN110": # Neni to kanbanovka.
            typ = "M" 
        else:
            typ= "P" # Je to kanbanovka. 
    elif (parameters.get(item).get("nakupci") == "U5004258") and (parameters.get(item).get("supplier") == "0"): # Stanjur bez nastavenych purchase dat → na routingy.
        typ = "M"
    elif (parameters.get(item).get("nakupci") == "PZP001") and (parameters.get(item).get("supplier") == "0"): # Generic purchaser bez nastavenych purchase dat → na routingy.
        typ = "M"    
    elif parameters.get(item).get("typ") == "Manufactured":
        typ = "M"
    elif (parameters.get(item).get("typ") == "Purchased") and (parameters.get(item).get("supplier") != "0") and (parameters.get(item).get("nakupci") != "0"):          
        typ = "P"
    # konkretni vyjimky itemy:
    elif item == "222516-15":
        typ = "P"
    else:
        typ = "N/A"
    return typ


def je_to_man_placard(item, parameters):
    nazev = parameters.get(item).get("description")
    man_placard = "placar"
    instalace   = "installation" 
    if (man_placard.upper() in nazev.upper()) and (parameters.get(item).get("typ") == "Manufactured"):
        if instalace.upper() not in nazev.upper():
            test = True
        else:
            test = False
    else:
        test = False
    return test


def level_pur_itemu(line, parameters):
    vyrabenych_dilu_v_lince = 0
    for item in line:
        typ = item_typ(item, parameters)
        if typ == "M":
            vyrabenych_dilu_v_lince += 1
        else:
            break
    if len(line) != vyrabenych_dilu_v_lince:
        prvni_nakupovany_poddil = line[vyrabenych_dilu_v_lince]
        level_prvniho_nakupovaneho_poddilu_linky = vyrabenych_dilu_v_lince + 1
    else:
        prvni_nakupovany_poddil = 0
        level_prvniho_nakupovaneho_poddilu_linky = 0
    # print("Prvni nakupovany poddil linky " + str(prvni_nakupovany_poddil))
    # print("Level prvniho nakupovaneho poddilu linky " + str(level_prvniho_nakupovaneho_poddilu_linky)) 
    return [prvni_nakupovany_poddil, level_prvniho_nakupovaneho_poddilu_linky, vyrabenych_dilu_v_lince]


def routing_lt(line, parameters, missing_lts): # Urceni MAN LT podle kolik je tam projektovych dilu a dohledani jejich Routing LT.
    chyby_routing_lt = []
    vsechny_chybejici_routingy_linky = []
    vsechny_neplatne_routingy = []
    vsechny_spatne_item_typ = []
    vsechny_expired_date = []
    vsechny_use_polozky = []

    vyrabeny_lt_linky = 0
    manufactured_placard = False

    for item in line:
        typ = item_typ(item, parameters)
        # print("item typ " + str(typ))
        if typ == "M": # Jedna se o MAN dil.
            if parameters.get(item).get("exdate") == "19-JAN-38": # S platnym expiry date.     
                manufactured_placard = je_to_man_placard(item, parameters)
                if manufactured_placard == True:
                    print(f'Vyrabeny item {item} v lince {line} je Manufactured PLACARD.')
                    vyrabeny_lt_linky += 2
                elif (parameters.get(item).get("description")[0:4] == "USE ") or (" USE " in parameters.get(item).get("description")):
                    print(f'Vyrabeny item {item} v lince {line} je USE polozka. Tato linka se nepocita.')
                    vsechny_use_polozky.append(item)
                elif parameters.get(item).get("phantom") == "Yes":
                    print(f'Vyrabeny item {item} v lince {line} je nastaven jako PHANTOM.')
                    vyrabeny_lt_linky += 0
                elif parameters.get(item).get("routinglt") != 0: # Dohledani routingu.
                    if parameters.get(item).get("routinglt") > 1999:
                        print(f'item {item} ma neplatny routing v LN. Tato linka se nepocita.')
                        vsechny_neplatne_routingy.append(item) 
                    else:
                        vyrabeny_lt_linky += parameters.get(item).get("routinglt")
                        vyrabeny_lt_linky += 3 # Pripocitava jeste 3 dni na kazdy vyrabeny dil na sestaveni.
                elif parameters.get(item).get("standardroutinglt") != 0: # pripadne standard routingu.
                    if parameters.get(item).get("standardroutinglt") > 1999:
                        print(f'item {item} ma neplatny Standard routing v LN. Tato linka se nepocita.')
                        vsechny_neplatne_routingy.append(item)
                    else:
                        vyrabeny_lt_linky += parameters.get(item).get("standardroutinglt")
                        vyrabeny_lt_linky += 3
                elif missing_lts !=0: # Pokud je nastaveno hromadne doplneni chybejicich routingu, bere hromadne nastavenu jako routing itemu.
                    if item not in vsechny_chybejici_routingy_linky:
                        vsechny_chybejici_routingy_linky.append(item)
                    vyrabeny_lt_linky += missing_lts
                    vyrabeny_lt_linky += 3                
                else: # Rucni zadani routingu, pokud ho to nenaslo v CQ exportu.
                    if parameters.get(item).get("supplier") == "I00000008": # Informace, zda se jedna o Lamph. dil.
                        lamphun = "(Lamphun)"
                    else:
                        lamphun = "" 
                    if item[0:3] == "PMP":
                        pn = item[9:len(item)]
                        project = item[0:9]
                    else:
                        pn = item
                        project = ""                        
                    print("Pozor! vyrabeny item " + str(project)+" "+str(pn)+str(lamphun) + " v lince " + str(line) + " nema zadny Routing LT. Bud nema routing, nebo routing neni nacteny v CQ.")                    
                    if item not in vsechny_chybejici_routingy_linky:
                        vsechny_chybejici_routingy_linky.append(item)
                    while parameters.get(item).get("routinglt") == 0:
                        try:
                            chybejici_routing_lt = int(input("Je treba dopsat do zdrojoveho souboru k prvnimu vyskutu itemu " + str(item) + ", nebo zadej zde:\n"))
                            if chybejici_routing_lt > 0:
                                parameters[item]["routinglt"] = chybejici_routing_lt  # Updatuje parametry itemu o chybejici udaj z inputu usera.
                                vyrabeny_lt_linky += chybejici_routing_lt
                                vyrabeny_lt_linky += 3
                                print("LT updatovan: "+str(parameters.get(item).get("routinglt")))
                                break
                            else: 
                                print("LT musi byt kladne cele cislo (pocet dni).")                                        
                        except ValueError:
                            print("LT musi byt cislo (pocet dni).")  
            else:
                print("Tato linka: " + str(line) + " se nepocita - vyrabeny dil " + str(item) + " nema platny Expiry date.\n")
                vsechny_expired_date.append(item)
                break
        elif typ == "N/A": # Neumime urcit typ itemu.
            # print(f'Nelze urcit typ itemu {item} v lince {line}. Tato linka se nepocita.')
            vsechny_spatne_item_typ.append(item)         
        else:
            break
    # print("Vyrabent LT linky " + str(vyrabeny_lt_linky))   
    # print(seznam_itemu_s_chybejicim_routingem)   
    chyby_routing_lt.append(vsechny_chybejici_routingy_linky)
    chyby_routing_lt.append(vsechny_neplatne_routingy)
    chyby_routing_lt.append(vsechny_spatne_item_typ)
    chyby_routing_lt.append(vsechny_expired_date)
    chyby_routing_lt.append(vsechny_use_polozky)

    # print(f'chyby vyrabenych dilu linky {line} {chyby_routing_lt}')
    return [vyrabeny_lt_linky, chyby_routing_lt]


def purchased_lt(line, parameters): # Najde prvni nakupovany dil pod poslednim vyrabenym dilem, overi jeho platnost a ziska jeho PUR LT.
    # print(f'resena linka PLTL: {line}')
    chyby_purchased_lt = []
    nakupovane_dily_bez_purchase_dat = []
    nakupovane_dily_bez_lt = []
    nakupovane_expired_date = []
    nakupovane_use_polozky = []

    nakupovany_lt_linky = 0
    nejnizsi_nakupovany_dil_v_lince = level_pur_itemu(line, parameters)[0]
    # print(nejnizsi_nakupovany_dil_v_lince)

    if nejnizsi_nakupovany_dil_v_lince == 0:
        print("Linka " + str(line) + " konci vyrabenym dilem s prazdnym kusovnikem. Bud se jedna o Fantom, MAN PLACARD, nebo Manufactured dil, ktery se pouze nakupuje.\n")
    else:
        if parameters.get(nejnizsi_nakupovany_dil_v_lince).get("exdate") == "19-JAN-38":            
            # konkretni vyjimky itemy:
            if nejnizsi_nakupovany_dil_v_lince == "222516-15":
                print(f' Nakupovany poddil {nejnizsi_nakupovany_dil_v_lince} v lince {line} je spatne nastaven v LN (ma byt P15/350S). LT natvrdo prirazen 60 dni.')
                nakupovany_lt_linky = 60            
            elif item_typ(nejnizsi_nakupovany_dil_v_lince, parameters) == "N/A":
                nakupovany_lt_linky = 0
            elif item_typ(nejnizsi_nakupovany_dil_v_lince, parameters) == "panel":
                nakupovany_lt_linky = 0
            elif (parameters.get(nejnizsi_nakupovany_dil_v_lince).get("description")[0:4] == "USE ") or (" USE " in parameters.get(nejnizsi_nakupovany_dil_v_lince).get("description")):
                print(f'Tato linka: {line} se nepocita! Nakupovany dil {nejnizsi_nakupovany_dil_v_lince} je USE polozka.')
                nakupovane_use_polozky.append(nejnizsi_nakupovany_dil_v_lince)    
            elif (parameters.get(nejnizsi_nakupovany_dil_v_lince).get("supplier") == "0") or (parameters.get(nejnizsi_nakupovany_dil_v_lince).get("nakupci") == "0"):
                print(f'Tato linka: {line} se nepocita! Nakupovany dil {nejnizsi_nakupovany_dil_v_lince} nema v purchase datech vyplneny dodavatele nebo nakupciho. Je nutno poptat v nakupu.')
                nakupovane_dily_bez_purchase_dat.append(nejnizsi_nakupovany_dil_v_lince)
            elif parameters.get(nejnizsi_nakupovany_dil_v_lince).get("phantom") == "Yes":
                print("Nakupovany item " + str(nejnizsi_nakupovany_dil_v_lince) + " v lince " + str(line) + " je nastaven jako Phantom")
                nakupovany_lt_linky = 0
            elif parameters.get(nejnizsi_nakupovany_dil_v_lince).get("warehouse") == "PZN110":
                nakupovany_lt_linky = 0
            elif (nejnizsi_nakupovany_dil_v_lince[0] == "3" or nejnizsi_nakupovany_dil_v_lince[0] == "9") and parameters.get(nejnizsi_nakupovany_dil_v_lince).get("safetystock") != 0:  # Overeni zda se jedna o surovy material se Safety stockem â†’ polud ano, LT 0.
                nakupovany_lt_linky = 0
            elif parameters.get(nejnizsi_nakupovany_dil_v_lince).get("supplytime") != 0:
                nakupovany_lt_linky = parameters.get(nejnizsi_nakupovany_dil_v_lince).get("supplytime")
            else:
                nakupovane_dily_bez_lt.append(nejnizsi_nakupovany_dil_v_lince)
                while True:
                    print("Pozor! Nakupovany dil " + str(nejnizsi_nakupovany_dil_v_lince) + " v lince " + str(line) + " ma v Purchase datech 0 LT!")
                    try:
                        chybejici_pur_lt = int(input("Dopln PUR LT pro " + str(nejnizsi_nakupovany_dil_v_lince) + " do txt. souboru, nebo zdej zde:\n"))
                        if chybejici_pur_lt > 0:
                            parameters[nejnizsi_nakupovany_dil_v_lince]["supplytime"] = chybejici_pur_lt  # Updatuje item o chybejici udaj z inputu usera.
                            nakupovany_lt_linky += chybejici_pur_lt
                            print("PUR LT updatovan.")
                            break
                        else:
                            print("PUR LT musi byt kladne cele cislo (pocet dni).")
                    except ValueError:
                        print("LT musi byt kladne cele cislo (pocet dni).")
                        continue
        elif parameters.get(line[line.index(nejnizsi_nakupovany_dil_v_lince)-1]).get("safetystock") != 0:
            print(f'Item {line[line.index(nejnizsi_nakupovany_dil_v_lince)-1]} (safety stock {parameters.get(line[line.index(nejnizsi_nakupovany_dil_v_lince)-1]).get("safetystock")}) ma pod sebou material {nejnizsi_nakupovany_dil_v_lince} s neplatnym expiry date.')
            nakupovany_lt_linky = 0            
        elif je_to_man_placard(line[line.index(nejnizsi_nakupovany_dil_v_lince)-1], parameters):
            print(f'MAM PLACARD {line[line.index(nejnizsi_nakupovany_dil_v_lince)-1]} s nesmazanym kusovnikem')
            nakupovany_lt_linky = 0
        else:
            print("Tato linka: " + str(line) + " se nepocita - nakupovany dil " + str(nejnizsi_nakupovany_dil_v_lince) + " nema platny Expiry date.\n")
            nakupovane_expired_date.append(nejnizsi_nakupovany_dil_v_lince)

    chyby_purchased_lt.append(nakupovane_dily_bez_purchase_dat)
    chyby_purchased_lt.append(nakupovane_dily_bez_lt)
    chyby_purchased_lt.append(nakupovane_expired_date)
    chyby_purchased_lt.append(nakupovane_use_polozky)   

    # print(f'chyby nakupovanuch dilu linky {line} {chyby_purchased_lt}')
    return [nakupovany_lt_linky, chyby_purchased_lt]


def lt_linky(line, parameters, missing_lts): # Vrati vysledny LT itemu a odpovidajici linku itemu.
    chyby_linky = []
    routing_lt_output = routing_lt(line, parameters, missing_lts)
    purchase_lt_output = purchased_lt(line, parameters)
    
    error_chybejici_routingy_linky = routing_lt_output[1][0]
    error_neplatne_routingy_linky = routing_lt_output[1][1]
    error_type = routing_lt_output[1][2]
    error_expiry_date_rlt = routing_lt_output[1][3]
    error_use_polozky_rlt = routing_lt_output[1][4] 
    error_itemy_bez_purchase_dat = purchase_lt_output[1][0]
    error_itemy_bez_purchase_lt = purchase_lt_output[1][1]
    error_expiry_date_plt = purchase_lt_output[1][2]
    error_nakupovane_use_polozky = purchase_lt_output[1][3]

    lt_linky = routing_lt_output[0]+purchase_lt_output[0]

    chyby_linky.append(error_chybejici_routingy_linky)

    chyby_linky.append(error_neplatne_routingy_linky)
    if len(error_neplatne_routingy_linky) != 0:
       lt_linky = 0 
    chyby_linky.append(error_type)
    if len(error_type) != 0:
       lt_linky = 0 
    chyby_linky.append(error_expiry_date_rlt)
    if len(error_expiry_date_rlt) != 0:
       lt_linky = 0 
    chyby_linky.append(error_itemy_bez_purchase_dat)    
    if len(error_itemy_bez_purchase_dat) != 0:
       lt_linky = 0 
    chyby_linky.append(error_itemy_bez_purchase_lt)

    chyby_linky.append(error_expiry_date_plt)    
    if len(error_expiry_date_plt) != 0:
       lt_linky = 0
    chyby_linky.append(error_use_polozky_rlt)    
    if len(error_use_polozky_rlt) != 0:
       lt_linky = 0
    chyby_linky.append(error_nakupovane_use_polozky)    
    if len(error_nakupovane_use_polozky) != 0:
       lt_linky = 0

    # print(f'vsechny chyby dilu linky {line} {chyby_linky}')
    # print(lt_linky)
    return [lt_linky, chyby_linky]

   
def nejdelsi_linka(line, parameters):
    lvl_pur_itemu = level_pur_itemu(line, parameters)[2]
    if lvl_pur_itemu != 0:
        nejdelsi_linka = line[0:(lvl_pur_itemu + 1)]
    else:
        nejdelsi_linka = line
    return nejdelsi_linka
  
      
def vysledek_itemu(nejdelsi_linka, parameters, vrchol, max_lt_itemu, missing_lts, vrchol_chyby): # sestaveni nejdelsi linky a jejiho LT. 
    lt_nejdelsi_linky = []
    vysledek_itemu = ""
    # print(f'MLT i {max_lt_itemu}')
    if max_lt_itemu == 0:
        print(f'U zadne linky itemu {vrchol} nebylo mozne spocitat platny LT.\n')
        sales_lt_itemu = "N/A"
        vysledek_itemu = f'U zadne linky itemu {vrchol} nebylo mozne spocitat platny LT.'
    elif (len(vrchol_chyby) == 1 and "error3" in vrchol_chyby) or len(vrchol_chyby) == 0:
        if je_to_man_placard(vrchol, parameters) == True:
            sales_lt_itemu = 14       
        else:
            sales_lt_itemu = int((max_lt_itemu/5+2)*7)
            if sales_lt_itemu <22:
                sales_lt_itemu = 22   
        for item in nejdelsi_linka:
            typ = item_typ(item, parameters)
            if typ == "M": # and (parameters.get(item).get("warehouse") != "PZN110"):
                if parameters.get(item).get("phantom") == "Yes":
                    lt_itemu = "MAN LT: " + "0 (phantom)" 
                elif parameters[item].get("routinglt") != 0:
                    lt_itemu = "MAN LT: " + str(parameters[item].get("routinglt"))
                elif parameters[item].get("standardroutinglt") != 0:
                    lt_itemu = "MAN LT: " + str(parameters[item].get("standardroutinglt"))
                elif missing_lts != 0: 
                    lt_itemu = "MAN LT: " + str(missing_lts)+"(plosny routing)"
                else: 
                    lt_itemu = "MAN LT: N/A"   
                lt_nejdelsi_linky.append(lt_itemu)
            elif ((item[0] == "3") or (item[0] == "9")) and parameters.get(item).get("safetystock") != 0:
                lt_itemu = "PUR LT: " + str(parameters[item].get("supplytime")) + " (safety stock: " + str(parameters[item].get("safetystock")) + ")"
                lt_nejdelsi_linky.append(lt_itemu)
            else:
                lt_itemu = "PUR LT: " + str(parameters[item].get("supplytime"))
                lt_nejdelsi_linky.append(lt_itemu)
        if vrchol[0:3] == "PMP": # Jedna se o projektovy vrchol
            print("\nLinka s nejdelsim LT itemu "+ str(vrchol[0:10]) + " " + str(vrchol[9:len(vrchol)]) + ":\n" + str(nejdelsi_linka) +"\n" + str(lt_nejdelsi_linky))
        else: # Jedna se o anonymni vrchol
            print("\nLinka s nejdelsim LT itemu "+str(vrchol[0:len(vrchol)]) + ":\n" + str(nejdelsi_linka) +"\n" + str(lt_nejdelsi_linky))
        print("Production LT itemu " + str(nejdelsi_linka[0]) + ": " + str(max_lt_itemu) + " pracovnich dni.")
        print("Sales LT itemu " + str(nejdelsi_linka[0]) + ": " + str(sales_lt_itemu) + " kalendarnich dni.\n")
        if vrchol[0:3] == "PMP":
            vysledek_itemu = str(vrchol[0:9])+":"+str(vrchol[9:len(vrchol)])+" = "+str(int(sales_lt_itemu))
        else:
            if parameters.get(vrchol).get("supplier") == "I00000008":  
                vysledek_itemu = "Lamphun_a:"+str(vrchol[0:len(vrchol)])+" = "+str(int(sales_lt_itemu))
            elif parameters.get(vrchol).get("supplier") != "0" and parameters.get(vrchol).get("nakukpci") != "0":
                vysledek_itemu = "ciste_nak:"+str(vrchol[0:len(vrchol)])+" = "+str(int(sales_lt_itemu))             
            else :
                vysledek_itemu = "anonymni_:"+str(vrchol[0:len(vrchol)])+" = "+str(int(sales_lt_itemu))     
    elif len(vrchol_chyby) != 0:
        sales_lt_itemu = "N/A"
        if vrchol[0:3] == "PMP":
            vysledek_itemu = f'{vrchol[0:9]}:{vrchol[9:len(vrchol)]} = '
        else:
            if parameters.get(vrchol).get("supplier") == "I00000008":  
                vysledek_itemu = f'Lamphun_a:{vrchol[0:len(vrchol)]} = '
            else :
                vysledek_itemu = f'anonymni_:{vrchol[0:len(vrchol)]} = '              
        if "error1" in vrchol_chyby:
                # print('Neplatny routing na nekterem z dilu vrcholu.') KRITICKA chyba
                vysledek_itemu += f'[Neplatny/e routing/gy.]  ' 
        if "error2" in vrchol_chyby:
                # print('Neplatny typ itemu na nekterem z dilu vrcholu.') KRITICKA chyba
                vysledek_itemu += f'[Neplatny typ itemu.]  ' 
        if "error3" in vrchol_chyby:
                # print('Neplatny expiry date na nekterem z manufactured dilu vrcholu.')
                vysledek_itemu += f'[Manufactured dil/y s neplatnym expire date.]  '         
        if "error4" in vrchol_chyby:
                # print('Chybi purcahse data na nekterem z dilu vrcholu.') KRITICKA chyba
                vysledek_itemu += f'[Chybi purchase data na nakupovanem dilu.]  '
        if "error5" in vrchol_chyby:
                # print('Neplatny expiry date na nekterem z purchased dilu vrcholu.') KRITICKA chyba
                vysledek_itemu += f'[Purchased dil/y s neplatnym expire date.]  '
        if "error6" in vrchol_chyby:
                # print('Manufactured USE polozka vrcholu.') KRITICKA chyba
                vysledek_itemu += f'[Manufactured USE polozka vrcholu.]  '
        if "error7" in vrchol_chyby:
                # print('nakupovana USE polozka vrcholu.') KRITICKA chyba
                vysledek_itemu += f'[Nakupovana USE polozka vrcholu.]  '
    return vysledek_itemu

def chyby_linky(
    line,
    vrchol_chyby,
    lt_linky_output,
    vsechny_chybejici_routingy,
    vsechny_neplatne_routingy,
    vsechny_itemy_typ_error,
    vsechny_nakupovane_bez_puchase_dat,
    vsechny_nakupovane_bez_puchase_lt,
    vsechny_vyrabene_use_polozky,
    vsechny_nakupovane_use_polozky
):
            for item in lt_linky_output[1][0]:
                if item not in vsechny_chybejici_routingy:
                    vsechny_chybejici_routingy.append(item)
            for item in lt_linky_output[1][1]:
                if item not in vsechny_neplatne_routingy:
                   vsechny_neplatne_routingy.append(item)
            for item in lt_linky_output[1][2]:
                if item not in vsechny_itemy_typ_error:
                   vsechny_itemy_typ_error.append(item)
            for item in lt_linky_output[1][4]:
                if item not in vsechny_nakupovane_bez_puchase_dat:
                    vsechny_nakupovane_bez_puchase_dat.append(item)
            for item in lt_linky_output[1][5]:
                if item not in vsechny_nakupovane_bez_puchase_lt:    
                   vsechny_nakupovane_bez_puchase_lt.append(item)
            for item in lt_linky_output[1][7]:
                if item not in vsechny_vyrabene_use_polozky:    
                   vsechny_vyrabene_use_polozky.append(item)
            for item in lt_linky_output[1][8]:
                if item not in vsechny_nakupovane_use_polozky:
                    vsechny_nakupovane_use_polozky.append(item)                                      
            
            if len(lt_linky_output[1][1]) != 0:
                print(f'Pozor! itemy: {lt_linky_output[1][1]} v lince {line} maji neplatny routing. Tato linka se nepocita.')
                if 'error1' not in vrchol_chyby:
                    vrchol_chyby.append('error1')
            elif len(lt_linky_output[1][2]) != 0:
                print(f'Pozor! U itemu: {lt_linky_output[1][2]} v lince {line} nelze urcit typ (Manufactured / Purchased). Tato linka se nepocita.')
                if 'error2' not in vrchol_chyby:
                    vrchol_chyby.append('error2')
            elif len(lt_linky_output[1][3]) != 0:
                print(f'Pozor! Item: {lt_linky_output[1][3]} v lince {line} nema platny expiry date. Tato linka se nepocita.')
                if 'error3' not in vrchol_chyby:
                    vrchol_chyby.append('error3')
            elif len(lt_linky_output[1][4]) != 0:
                print(f'Pozor! nakupovane itemy: {lt_linky_output[1][4]} v lince {line} nemaji vyplneno suppliera nebo nakupciho. Tato linka se nepocita.')
                if 'error4' not in vrchol_chyby:
                    vrchol_chyby.append('error4')               
            elif len(lt_linky_output[1][6]) != 0:
                print(f'Pozor! nakupovane itemy: {lt_linky_output[1][6]} v lince {line} nemaji platny expiry date. Tato linka se nepocita.')
                if 'error5' not in vrchol_chyby:
                    vrchol_chyby.append('error5')
            elif len(lt_linky_output[1][7]) != 0:
                print(f'Pozor! vyrabene itemy: {lt_linky_output[1][7]} v lince {line} jsou USE polozky. Tato linka se nepocita.')
                if 'error6' not in vrchol_chyby:
                    vrchol_chyby.append('error6')
            elif len(lt_linky_output[1][8]) != 0:
                print(f'Pozor! nakupovane itemy: {lt_linky_output[1][8]} v lince {line} jsou USE polozky. Tato linka se nepocita.')
                if 'error7' not in vrchol_chyby:
                    vrchol_chyby.append('error7')

def linka_k_zaplanovani(line, parameters):
    # print(line)
    vse_v_lince = []
    for item in line:
        typ = item_typ(item, parameters)
        # print(typ)
        if typ == "M":
            if parameters.get(item).get("exdate") == "19-JAN-38":
                if parameters.get(item).get("phantom") == "Yes":
                    if item [0:3] == "PMP":
                        item_to_plan = f'{item}(PHANTOM)'
                    else:
                        item_to_plan = f'PROJEKT_{item}(PHANTOM)'
                elif line.index(item)+1 != len(line):
                    if parameters.get(line[line.index(item)+1]).get("exdate") != "19-JAN-38":
                        if parameters.get(item).get("safetystock") != 0:
                            if item [0:3] == "PMP":
                                item_to_plan = f'{item[10:len(item)]} manufactued dil s neplatnym materialem. Pouze nakoupit.(ss {parameters.get(item).get("safetystock")})'
                            else:
                                item_to_plan = f'{item} manufactued dil s neplatnym materialem. Pouze nakoupit.(ss {parameters.get(item).get("safetystock")})'                 
                    elif item [0:3] == "PMP":
                        item_to_plan = item
                    else:
                        item_to_plan = f'PROJEKT_{item}'
                elif item [0:3] == "PMP":
                    item_to_plan = item
                else:
                    item_to_plan = f'PROJEKT_{item}'
            else:   
                return (vse_v_lince)      
        elif typ == "P":
            if parameters.get(item).get("exdate") == "19-JAN-38":
                item_to_plan = item
            elif item_typ(line[line.index(item)-1], parameters) == "M":
                if parameters.get(line[line.index(item)-1]).get("safetystock") != 0:
                    item_to_plan = f'{item} Material s neplatnym expiry date. Vyrabeny naddil {[line.index(item)-1]} pouze nakoupit(ss {parameters.get(item).get("safetystock")})'
                else:
                    item_to_plan = f'ERROR material {item} ma neplatny expiry date!'
            elif parameters.get(item).get("exdate") != "19-JAN-38":
                item_to_plan = ""
        elif typ == "panel":
            item_to_plan = item            
        elif typ == "N/A":
            item_to_plan = f'ErrorType {item}'
        vse_v_lince.append(item_to_plan)
    return(vse_v_lince)


def zaplanovani_do_excelu(vse_k_zaplanovani):
    wb = excel.load_workbook('C:\\Users\\Ondrej.rott\\Documents\\Python\\Pracovni\\to_plan.xlsx')
    sheet1 = wb["K zaplanovani do LN"]
    radek_itemu_to_plan = 2
    sloupec_itemu_to_plan = 1
   
    for row in sheet1:
        sheet1.delete_rows(2,sheet1.max_row)
    for linka in vse_k_zaplanovani:
        if vse_k_zaplanovani.index(linka) == 0: # prvni linka k zaplanovani
            current_vrchol = linka[0]
            for item in linka:       
                cell_to_plan = sheet1.cell(radek_itemu_to_plan,sloupec_itemu_to_plan)
                cell_to_plan.value = item
                sloupec_itemu_to_plan += 1
            radek_itemu_to_plan += 1
            sloupec_itemu_to_plan = 1        
        else: # nejedna se o prvni linku k zaplanovani
            for item in linka:
                try:
                    if  linka.index(item) == 0: # pro prvni item v lince je test jiny nez pro zbytek 
                        if item == current_vrchol:
                            cell_to_plan = sheet1.cell(radek_itemu_to_plan,sloupec_itemu_to_plan)
                            cell_to_plan.value = None
                            sloupec_itemu_to_plan += 1
                        else:
                            current_vrchol = item
                            cell_to_plan = sheet1.cell(radek_itemu_to_plan,sloupec_itemu_to_plan)
                            cell_to_plan.value = item
                            sloupec_itemu_to_plan += 1
                    # elif not(len(linka) > len(vse_k_zaplanovani[vse_k_zaplanovani.index(linka)-1])): # predchozi linka nesmi byt kratsi nez tato. (muze se jednat o pokracovani predchoziho poddilu)
                    elif item == vse_k_zaplanovani[vse_k_zaplanovani.index(linka)-1][linka.index(item)+linka[0:sloupec_itemu_to_plan-1].count(item)]: #pokud je item stejny jako ten nad nim
                        if sheet1.cell(radek_itemu_to_plan, sloupec_itemu_to_plan -1).value == None: # pokud bunka v excelu vlevo od itemu je prazdna (jedna se o poddil predchoyiho itemu) 
                            cell_to_plan = sheet1.cell(radek_itemu_to_plan,sloupec_itemu_to_plan)
                            cell_to_plan.value = None # None → pokracovni dilu nad timto
                            sloupec_itemu_to_plan += 1
                        else: 
                            cell_to_plan = sheet1.cell(radek_itemu_to_plan,sloupec_itemu_to_plan)
                            cell_to_plan.value = item # jedna se o stejny dil jako nad nim, ale pod jinym vrcholem → zaplanovat
                            sloupec_itemu_to_plan += 1
                    else:
                        cell_to_plan = sheet1.cell(radek_itemu_to_plan,sloupec_itemu_to_plan)
                        cell_to_plan.value = item # jedna se o stejny dil jako nad nim, ale pod jinym vrcholem → zaplanovat
                        sloupec_itemu_to_plan += 1
                except:
                    cell_to_plan = sheet1.cell(radek_itemu_to_plan,sloupec_itemu_to_plan)
                    cell_to_plan.value = item # jedna se o stejny dil jako nad nim, ale pod jinym vrcholem → zaplanovat
                    sloupec_itemu_to_plan += 1
            radek_itemu_to_plan +=1
            sloupec_itemu_to_plan = 1
    wb.save("C:\\Users\\Ondrej.rott\\Documents\\Python\\Pracovni\\to_plan.xlsx")