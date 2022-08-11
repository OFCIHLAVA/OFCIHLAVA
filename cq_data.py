import datetime
from http.client import LineTooLong
from itertools import count
from ntpath import join
from tabnanny import check
import openpyxl as excel

from datetime import date, timedelta
import time
from operator import index
import warnings


def data_import(cq_export_cesta): # Nacteni dat z CQ reportu.
    with open(cq_export_cesta,'r') as input_file:
        data = input_file.readlines()
        input_file.close()
    return data

def data_headings(data): # Vytvoreni zahlavi sloupcu z reportu.
    for line in data:
        if line[2:5] == "***": # Identifikator linky se zahlavim z cq reportu.
            data_headings = [pole.strip() for pole in line.split("|")]
            data_headings[0] = "Item line in LN" # Prepsani "***" v prvnim sloupci, kde bude pozdeji doplneno poradi linky v cq reportu.
    return data_headings

def import_data_cleaning(data): # Ocisteni dat a nahrazeni prazdnych poli za "0".
    cl_data = list()    
    for line in data: # Ocisteni dat a 
        if line[0] == "|":
            linka = [pole.strip() for pole in line.split("|")] # Ocisteni dat.
            for pole in linka: # nahrazeni praydnych poli za "0".
                if len(pole) == 0:
                    if linka.index(pole) == 0:
                        linka.remove(pole)
                    else:
                        linka[linka.index(pole)] = "0"
            cl_data.append(linka)
            data = cl_data
    return data

def add_line_id_to_data(data, data_headings): # Pridani ID linky pro jednotlive vrcholy do cq dat jako int na pozici 0 kazde linky. Zvlast resen sklad PZN100 a PZN105.
    pocet_linek_vrcholu_PZN100 = {}
    pocet_linek_vrcholu_PZN105 = {}
    
    for heading in data_headings: # Nalezeni indexu sloupcu skladu a vrcholu v cq importu.
        if "CLUSTER" in heading.upper():
            sklad_index = data_headings.index(heading)
        if heading.upper() == "ITEM":
            vrchol_index = data_headings.index(heading)
    
    for line in data:
        sklad = line[sklad_index-1]
        vrchol = line[vrchol_index-1]

        # PZN100
        if sklad.upper() == "PZN100":
            if vrchol not in pocet_linek_vrcholu_PZN100:
                pocet_linek_vrcholu_PZN100[vrchol] = 0 
            line.insert(0, pocet_linek_vrcholu_PZN100.get(vrchol)+1) # Vlozeni sloupce ID linky pred stavavjici data. (rozsireni dat o jeden sloupec)
            # print(line)
            pocet_linek_vrcholu_PZN100[vrchol] +=1

        # PZN105
        elif sklad.upper() == "PZ5":
            # print(f"PZN105")
            if vrchol not in pocet_linek_vrcholu_PZN105:
                pocet_linek_vrcholu_PZN105[vrchol] = 0 
            line.insert(0, pocet_linek_vrcholu_PZN105.get(vrchol)+1)
            pocet_linek_vrcholu_PZN105[vrchol] +=1
    return data

def data_date_formating(data, data_headings): # Prevedeni pole datumu na date format.
    for heading in data_headings:
        if "DATE" in heading.upper():
            date_index = data_headings.index(heading)
            break

    for line in data: # Prevedeni pole datumu na date format.
        den, mesic, rok = line[date_index].split("/")
        # print(rok, mesic, den)
        datum = datetime.date(int(rok), int(mesic), int(den))
        line[date_index] = datum
    return data

def order_plan_database_pzn100(data, headings): # Vytvoreni dictionary PZN100 jednotlivych itemu s jejich linkamy a daty z reportu.
    database_pzn100_order_plan_dictionary = dict() # Vsechny vrcholy z PZN100 z reportu : jejich linky.
    data_line_dict_100 = {} # Jednotlive linky vrcholu z PZN100 : data k linkam.
    data_dict_100 = {} # Samotna data na linkach z PZN100.
    for heading in headings:
        if heading.upper() == "ITEM":
            vrchol_index = headings.index(heading)
        if heading.upper() == "CLUSTER":
            sklad_index = headings.index(heading)
    vrchol = data[0][vrchol_index]
    
    for line in data:        
        # PZN100:
        if line[sklad_index] == "PZN100":       
            if line[vrchol_index] == vrchol: # Pokud se jedna o pokracovani stejneho vrcholu.
                for data_field in line:
                    data_dict_100[headings[line.index(data_field)]] = data_field # Sestaveni dat z linky jako dict s jmeny sloupcu zahlavi reportu jako klic.
                data_line_dict_100[line[0]] = data_dict_100 # Sestaveni jednotlivych linek jako dict cislo linky vrcholu : data v lince vyse.
                data_dict_100 = {} # Reset dat pro novou linku.      
            else: # Pokud se jedna o novy vrchol.
                database_pzn100_order_plan_dictionary[vrchol] = data_line_dict_100 # Sestaveni linek predchoziho vrcholu jako dict vrchol : jeho linky s daty viz. vyse.
                
                # Reset pro novy vrchol.
                vrchol = line[vrchol_index] 
                data_line_dict_100 ={}
                data_dict_100 = {}
                
                # Opakovani viz vyse pro novy vrchol.
                for data_field in line:
                    data_dict_100[headings[line.index(data_field)]] = data_field    
                data_line_dict_100[line[0]] = data_dict_100
                data_dict_100 = {}
            database_pzn100_order_plan_dictionary[vrchol] = data_line_dict_100
    return database_pzn100_order_plan_dictionary

def order_plan_database_pzn105(data, headings): # Vytvoreni dictionary PZN105 jednotlivych itemu s jejich linkamy a daty z reportu.
    database_pzn105_order_plan_dictionary = dict() # Vsechny vrcholy z PZN105 z reportu : jejich linky.
    data_line_dict_105 = {} # Jednotlive linky vrcholu z PZN105 : data k linkam.
    data_dict_105 = {} # Samotna data na linkach z PZN105.
    for heading in headings:
        if heading.upper() == "ITEM":
            vrchol_index = headings.index(heading)
        if heading.upper() == "CLUSTER":
            sklad_index = headings.index(heading)
    vrchol = data[0][vrchol_index]

    for line in data:        
        # PZN105:
        if line[sklad_index] == "PZ5":       
            if line[vrchol_index] == vrchol: # Pokud se jedna o pokracovani stejneho vrcholu.
                for data_field in line:
                    data_dict_105[headings[line.index(data_field)]] = data_field # Sestaveni dat z linky jako dict s jmeny sloupcu zahlavi reportu jako klic.
                data_line_dict_105[line[0]] = data_dict_105 # Sestaveni jednotlivych linek jako dict cislo linky vrcholu : data v lince vyse.
                data_dict_105 = {} # Reset dat pro novou linku.      
            else: # Pokud se jedna o novy vrchol.
                database_pzn105_order_plan_dictionary[vrchol] = data_line_dict_105 # Sestaveni linek predchoziho vrcholu jako dict vrchol : jeho linky s daty viz. vyse.
                
                # Reset pro novy vrchol.
                vrchol = line[vrchol_index] 
                data_line_dict_105 ={}
                data_dict_105 = {}
                
                # Opakovani viz vyse pro novy vrchol.
                for data_field in line:
                    data_dict_105[headings[line.index(data_field)]] = data_field    
                data_line_dict_105[line[0]] = data_dict_105
                data_dict_105 = {}
            database_pzn105_order_plan_dictionary[vrchol] = data_line_dict_105
    return database_pzn105_order_plan_dictionary    



    # elif line[1] == "PZ5":  
    #     if line[2] == vrchol: # Pokud se jedna o pokracovani stejneho vrcholu.
    #         for data_field in line:
    #             data_dict_105[s_zahlavi[line.index(data_field)]] = data_field # Sestaveni dat z linky jako dict s jmeny sloupcu zahlavi reportu jako klic.
    #         data_line_dict_105[line[0]] = data_dict_105 # Sestaveni jednotlivych linek jako dict cislo linky vrcholu : data v lince vyse.
    #         data_dict_105 = {} # Reset dat pro novou linku.      
    #     else: # Pokud se jedna o novy vrchol.
    #         database_pzn105_order_plan_dictionary[vrchol] = data_line_dict_105 # Sestaveni linek predchoziho vrcholu jako dict vrchol : jeho linky s daty viz. vyse.
    #         
    #         # Reset pro novy vrchol.
    #         vrchol = line[2] 
    #         data_line_dict_105 ={}
    #         data_dict_105 = {}
    #         
    #         # Opakovani viz vyse pro novy vrchol.
    #         for data_field in line:
    #             data_dict_105[s_zahlavi[line.index(data_field)]] = data_field    
    #         data_line_dict_105[line[0]] = data_dict_105
    #         data_dict_105 = {}
    #     database_pzn105_order_plan_dictionary[vrchol] = data_line_dict_105
# 
# # print(f'PZN 100 databaze') 
# # print(len(database_pzn100_order_plan_dictionary))
# # for key in database_pzn100_order_plan_dictionary:
# #     print(key)
# #     print(database_pzn100_order_plan_dictionary[key])
# # print(f'PZN 105 databaze') 
# # print(len(database_pzn105_order_plan_dictionary))
# # for key in database_pzn105_order_plan_dictionary:
# #     print(key)
# #     print(database_pzn105_order_plan_dictionary[key])
# 
#### Ziskani dat shoratage linek z Master planu

#9999 wb_master_plan = excel.load_workbook("Y:\\Departments\\Sales and Marketing\\Shared\\Aftersales\\SCCZ-AS-F003 Aftersales Master Plan.xlsm")
#9999 sheet1_master_plan = wb_master_plan.worksheets[0]
#9999 
#9999 # najit spravne sloupce 
#9999 # Sales order, Item, Description, PDD, CRD, SPD, Ordered Qty, Availability
#9999 
#9999 # Ziskani nazvu sloupcu.
#9999 zahlavi_master_planu = []
#9999 for c in range(1,min(sheet1_master_plan.max_column+1, 16384)):
#9999     column_name = sheet1_master_plan.cell(1, c).value
#9999     if column_name not in zahlavi_master_planu and column_name:
#9999         zahlavi_master_planu.append(column_name)
#9999 # print(zahlavi_master_planu)
#9999 
#9999 for nazev in zahlavi_master_planu:
#9999     # Najit sloupec Sales order.
#9999     if nazev.upper() == "SALES ORDER":
#9999         sales_order_sloupec_index = zahlavi_master_planu.index(nazev)
#9999     # Najit Project.
#9999     elif nazev.upper() == "PROJECT":
#9999         project_sloupec_index = zahlavi_master_planu.index(nazev)
#9999     # Najit sloupec Item.
#9999     elif nazev.upper() == "ITEM":
#9999         item_sloupec_index = zahlavi_master_planu.index(nazev)
#9999     # Najit sloupec Description.
#9999     elif nazev.upper() == "DESCRIPTION":
#9999         description_sloupec_index = zahlavi_master_planu.index(nazev)
#9999     # Najit sloupec PDD.
#9999     elif nazev.upper() == "PLANNED DELIVERY DATE":
#9999         pdd_sloupec_index = zahlavi_master_planu.index(nazev)
#9999     # Najit sloupec CRD.
#9999     elif nazev.upper() == "CUSTOMER REQ DATE":
#9999         crd_sloupec_index = zahlavi_master_planu.index(nazev)
#9999     # Najit sloupec SPD.
#9999     elif nazev.upper() == "SUPPLIER PROMISED DATE":
#9999         spd_sloupec_index = zahlavi_master_planu.index(nazev)
#9999     # Najit sloupec Ordered Qty.
#9999     elif nazev.upper() == "ORDERED QTY":
#9999         ordered_qty_sloupec_index = zahlavi_master_planu.index(nazev)
#9999     # Najit sloupec Availability.
#9999     elif nazev.upper() == "AVAILABILITY":
#9999         availability_sloupec_index = zahlavi_master_planu.index(nazev)
#9999 
#9999 # Urceni datumu, do kdy proverovat linky.
#9999 neplatne_datum_ln = date(1970, 1, 1)
#9999 dnes = date.today()
#9999 proverit_pristich_dni = 5
#9999 posun_dnu = timedelta(proverit_pristich_dni)
#9999 proverit_do_datumu = dnes + posun_dnu
#9999 # print(f'Dnes {dnes}')
#9999 # print(f'Proverit do {proverit_do_datumu}')
#9999 
#9999 datumy_sloupce = [pdd_sloupec_index, crd_sloupec_index, spd_sloupec_index]
#9999 # print(f'datumy sloupce indexy PDD, CRD, SPD {datumy_sloupce}')
#9999 
#9999 # Vytazeni dotcenych linek z Master planu.
#9999 itemy_k_prevedeni = list()
#9999 
#9999 # Pridani zahlavi do vysledku.
#9999 zahlavi = list()
#9999 
#9999 zahlavi.append(sheet1_master_plan.cell(1, item_sloupec_index+1).value)
#9999 zahlavi.append("Master plan row")
#9999 zahlavi.append(sheet1_master_plan.cell(1, description_sloupec_index+1).value)
#9999 zahlavi.append(sheet1_master_plan.cell(1, sales_order_sloupec_index+1).value)
#9999 zahlavi.append(sheet1_master_plan.cell(1, pdd_sloupec_index+1).value)
#9999 zahlavi.append(sheet1_master_plan.cell(1, crd_sloupec_index+1).value)
#9999 zahlavi.append(sheet1_master_plan.cell(1, spd_sloupec_index+1).value)
#9999 zahlavi.append("Buyer")
#9999 zahlavi.append(sheet1_master_plan.cell(1, ordered_qty_sloupec_index+1).value)
#9999 zahlavi.append("Sum req. Qty Item+PDD")
#9999 zahlavi.append("Planned available Qty on PZN105 at PDD")
#9999 zahlavi.append("Planned available Qty on PZN100 at PDD")
#9999 zahlavi.append("Already requested in Tabulka prevodu na PZN105")
#9999 zahlavi.append("PZN 105 Nejblizsi datum + mnozstvi: kdy PA >= 0.")
#9999 zahlavi.append("PZN 100 Nejblizsi datum + mnozstvi: kdy PA >= 0.")
#9999 
#9999 print(zahlavi)
#9999 
#9999 # print(sheet1_master_plan.max_row)
#9999 for row in range(2 ,min(1048576, sheet1_master_plan.max_row+1)):    
#9999     # print(sheet1_master_plan.cell(row, item_sloupec_index+1).value)
#9999     # 1. Kontrola Shortage linky.
#9999     if sheet1_master_plan.cell(row, availability_sloupec_index+1).value != None and sheet1_master_plan.cell(row, availability_sloupec_index+1).value.upper() == "SHORTAGE": 
#9999         # 2. Kontrola anonymni polozky.
#9999         # print(sheet1_master_plan.cell(row, project_sloupec_index+1).value)
#9999         if sheet1_master_plan.cell(row, project_sloupec_index+1).value == "":
#9999             # 3. Kontrola datumu.
#9999             for bunka in [sheet1_master_plan.cell(row, c+1).value.date() for c in datumy_sloupce]:
#9999                 if bunka != neplatne_datum_ln and bunka <= proverit_do_datumu and bunka >= dnes:
#9999                     # print(f'ZASAH!')
#9999                     # f = f'shortage na lince {row}'
#9999                     # check = [sheet1_master_plan.cell(row, c+1).value.date() for c in datumy_sloupce]
#9999                     # print(f'bunky k otestovani na datum 1 {check}')
#9999                     # print(f)
#9999                     list_linky = []
#9999                     list_linky.append(sheet1_master_plan.cell(row, item_sloupec_index+1).value)
#9999                     list_linky.append(row)
#9999                     list_linky.append(sheet1_master_plan.cell(row, description_sloupec_index+1).value)
#9999                     list_linky.append(sheet1_master_plan.cell(row, sales_order_sloupec_index+1).value)
#9999                     list_linky.append(sheet1_master_plan.cell(row, pdd_sloupec_index+1).value.strftime("%d/%m/%Y").replace("/", "."))
#9999                     list_linky.append(sheet1_master_plan.cell(row, crd_sloupec_index+1).value.strftime("%d/%m/%Y").replace("/", "."))
#9999                     list_linky.append(sheet1_master_plan.cell(row, spd_sloupec_index+1).value.strftime("%d/%m/%Y").replace("/", "."))
#9999                     # Pridani sloupce buyer
#9999                     if database_pzn100_order_plan_dictionary.get(sheet1_master_plan.cell(row, item_sloupec_index+1).value.strip()) or database_pzn105_order_plan_dictionary.get(sheet1_master_plan.cell(row, item_sloupec_index+1).value.strip()):
#9999                         buyer = database_pzn100_order_plan_dictionary.get(sheet1_master_plan.cell(row, item_sloupec_index+1).value.strip()).get(1).get("Buyer") if database_pzn100_order_plan_dictionary.get(sheet1_master_plan.cell(row, item_sloupec_index+1).value.strip()) else database_pzn105_order_plan_dictionary.get(sheet1_master_plan.cell(row, item_sloupec_index+1).value.strip()).get(1).get("Buyer")
#9999                     else:
#9999                         buyer = "0"
#9999                     list_linky.append(buyer)
#9999                     list_linky.append(sheet1_master_plan.cell(row, ordered_qty_sloupec_index+1).value)
#9999                     print(zahlavi)
#9999                     print(list_linky)
#9999                     itemy_k_prevedeni.append(list_linky)
#9999                     list_linky = list()          
#9999                     break       
#9999 
#9999 # Serazeni vysledku podle itemu a linky Master planu.
#9999 itemy_k_prevedeni.sort()
#9999 # print(len(itemy_k_prevedeni))
#9999 
#9999 # Doplneni zahlavi.
#9999 itemy_k_prevedeni.insert(0, zahlavi)
#9999 # print()
#9999 # print(f'Seznam PZN 105 shortage linek z Master planu.')
#9999 # for line in itemy_k_prevedeni:
#9999 #     print(line)
#9999 
#9999 # Pro kazdou linku vytazenou z Master planu prida na konec linky sumu pozadavku Qty v dany den pro dany vrchol a PDD linky  (suma vsech stejnych vrcholu v dane PDD).
#9999 for line in itemy_k_prevedeni:
#9999     if line[0] == "Item":
#9999         continue        
#9999     vrchol = line[0]
#9999     pdd = line[4]
#9999     requested_qty_vrchol_pdd = 0
#9999     for linka in itemy_k_prevedeni: # Projde vsechny linky a nacte pozadavky na Qty pro dane PDD datum pro dany vrchol.
#9999         if linka[0] == vrchol and linka[4] == pdd:
#9999             requested_qty_vrchol_pdd += linka[8]
#9999     line.append(float(requested_qty_vrchol_pdd)) # Pripoji udaj na konec linky.
#9999 print(f'SUM qty')
#9999 for line in itemy_k_prevedeni:
#9999     print(line)
#9999 
#9999 # Dotaz na Planned available daneho itemu v dane PDD podle linky. (realny stav / shortage na PZN105:)
#9999 for line in itemy_k_prevedeni:
#9999     order_type = ""
#9999     # print(line)    
#9999     vrchol = line[0]
#9999     if vrchol == "Item":
#9999         continue
#9999     pdd = line[4].split(".")
#9999     pdd = datetime.date(int(pdd[-1]), int(pdd[1]), int(pdd[0]))
#9999     vrchol_available_qty_PZN105 = 0
#9999     if database_pzn105_order_plan_dictionary.get(vrchol) != None:
#9999         for linka in database_pzn105_order_plan_dictionary.get(vrchol):
#9999             order_type = database_pzn105_order_plan_dictionary.get(vrchol).get(linka).get("Order type txt")
#9999             if order_type.upper() == "PLANNED PURCHASE ORDER":
#9999                 order_type = ""
#9999                 continue
#9999             # print(linka)
#9999             datum = database_pzn105_order_plan_dictionary.get(vrchol).get(linka).get("Date")            
#9999             if datum <= pdd:
#9999                 balance = database_pzn105_order_plan_dictionary.get(vrchol).get(linka).get("Transaction type txt")
#9999                 quantity = float(database_pzn105_order_plan_dictionary.get(vrchol).get(linka).get("Quantity"))                
#9999                 # print(quantity, type(quantity))
#9999                 if balance == "+ (Planned Receipt)":
#9999                     vrchol_available_qty_PZN105 += quantity
#9999                 elif balance == "- (Planned Issue)": 
#9999                     vrchol_available_qty_PZN105 -= quantity
#9999                 else:
#9999                     print('POZOR - ERROR v +/- balance na u itemu {vrchol} na lince linka.')
#9999     # print(f'Item {vrchol}, Planned available qty na PZN105 {vrchol_available_qty_PZN105} k datu {pdd}')
#9999     line.append(float(vrchol_available_qty_PZN105))
#9999 
#9999 
#9999 # Dotaz na Planned available daneho itemu v dane PDD podle linky. (realny stav / shortage na PZN100:)
#9999 for line in itemy_k_prevedeni:
#9999     order_type = ""
#9999     # print(line)    
#9999     vrchol = line[0]
#9999     if vrchol == "Item":
#9999         continue
#9999     pdd = line[4].split(".")
#9999     pdd = datetime.date(int(pdd[-1]), int(pdd[1]), int(pdd[0]))
#9999     vrchol_available_qty_PZN100 = 0
#9999     if database_pzn100_order_plan_dictionary.get(vrchol) != None:
#9999         for linka in database_pzn100_order_plan_dictionary.get(vrchol):
#9999             order_type = database_pzn100_order_plan_dictionary.get(vrchol).get(linka).get("Order type txt")
#9999             if order_type.upper() == "PLANNED PURCHASE ORDER":         
#9999                 order_type = ""
#9999                 continue            
#9999             # print(linka)
#9999             datum = database_pzn100_order_plan_dictionary.get(vrchol).get(linka).get("Date")
#9999             if datum <= pdd:
#9999                 balance = database_pzn100_order_plan_dictionary.get(vrchol).get(linka).get("Transaction type txt")
#9999                 quantity = float(database_pzn100_order_plan_dictionary.get(vrchol).get(linka).get("Quantity"))                
#9999                 # print(quantity, type(quantity))
#9999                 if balance == "+ (Planned Receipt)":
#9999                     vrchol_available_qty_PZN100 += quantity
#9999                 elif balance == "- (Planned Issue)": 
#9999                     vrchol_available_qty_PZN100 -= quantity
#9999                 else:
#9999                     print('POZOR - ERROR v +/- balance na u itemu {vrchol} na lince linka.')
#9999     # print(f'Item {vrchol}, Planned available qty na PZN100 {vrchol_available_qty_PZN100} k datu {pdd}')
#9999     line.append(float(vrchol_available_qty_PZN100))
#9999 
#9999 print(f'SUM PA PDD')
#9999 # for line in itemy_k_prevedeni:
#9999 #     print(line)
#9999 # 
#9999 # print()
#9999 # for line in itemy_k_prevedeni:
#9999 #     print(str(line[0]))
#9999 # 
#9999 # print()
#9999 # for line in itemy_k_prevedeni:
#9999 #     print(line)
#9999 
#9999 # Dotaz na tabulku prevodu, jestli uz je nejake mnozstvi zadane na prevod.
#9999 warnings.simplefilter(action='ignore', category=UserWarning)
#9999 wb_tabulka_prevodu = excel.load_workbook("Y:\\Departments\\Logistics\\Shared\WAREHOUSE\\After sales\\Prevody na PZN105_2022.xlsx")
#9999 
#9999 sheet1_tabulka_prevodu = wb_tabulka_prevodu.worksheets[0]
#9999 print("AHOAHOAHOJA")
#9999 
#9999 item_col = 2
#9999 qty_col = 3
#9999 from_whs_col = 7
#9999 to_whs_col = 8
#9999 date_created_col = 10
#9999 who_processed_col = 11
#9999 date_processed_col = 12
#9999 comment_col = 13
#9999 
#9999 for line in itemy_k_prevedeni:
#9999     # print(line)
#9999     vrchol = line[0]
#9999     if vrchol == "Item":
#9999         continue
#9999     vrchol_uz_zadano_k_prevodu = 0
#9999     # print(vrchol)
#9999     for row in range(7 ,min(1048576, sheet1_tabulka_prevodu.max_row+1)):
#9999         cell_item = sheet1_tabulka_prevodu.cell(row, item_col)
#9999         cell_who_processed = sheet1_tabulka_prevodu.cell(row, who_processed_col)
#9999         cell_created = sheet1_tabulka_prevodu.cell(row, date_created_col)
#9999         cell_qty = sheet1_tabulka_prevodu.cell(row, qty_col)
#9999         cell_from_whs = sheet1_tabulka_prevodu.cell(row, from_whs_col)
#9999         cell_to_whs = sheet1_tabulka_prevodu.cell(row, to_whs_col)
#9999 
#9999         if str(cell_who_processed.value).strip() == "None":
#9999             # print(row, cell_created.value, type(cell_created.value),sep="|")
#9999             # print(cell_created.value)
#9999             # print(type(cell_created.value))
#9999             
#9999             if str(type(cell_created.value)) == "<class 'str'>":
#9999                 try:
#9999                     # print(row)
#9999                     if "/" in cell_created.value[0:4]:
#9999                         splitted_date = cell_created.value.split("/")
#9999                         splitted_date = date(int(splitted_date[2][0:4]),int(splitted_date[1]),int(splitted_date[0]))
#9999                     if "." in cell_created.value[0:4]:
#9999                         splitted_date = cell_created.value.split(".")
#9999                         datum_created = date(int(splitted_date[2][0:4]),int(splitted_date[1]),int(splitted_date[0]))
#9999                 except:
#9999                     print(f'POZOR! Datum Kdy zadano v souboru Prevody na PZN105 na radce {row} je v nespravnem formatu. Nebylo mozno overit stav na teto radce.')
#9999                     continue                  
#9999             elif str(type(cell_created.value)) == "<class 'datetime.datetime'>":
#9999                 datum_created = cell_created.value.date()
#9999 
#9999             if datum_created > dnes-timedelta(4):
#9999 
#9999                 if str(cell_from_whs.value).upper() == "PZN100" and str(cell_to_whs.value).upper() == "PZN105":
#9999 
#9999                     if cell_item.value == vrchol:
#9999                         # print(row, {cell_item.value})
#9999                         # print(cell_created.value)
#9999                         # print(type(cell_created.value))
#9999                         vrchol_uz_zadano_k_prevodu += int(cell_qty.value) 
#9999 
#9999     line.append(vrchol_uz_zadano_k_prevodu)
#9999 
#9999 print(f'Uz zadano k prevodu')
#9999 for line in itemy_k_prevedeni:
#9999     print(line)
#9999 
#9999 
#9999 # Dotaz na nejblizsi datum, kdy bude planned available na skladech PZN105 + PZN100 pro dany item v lince alespon 0 nebo vyssi.
#9999 for line in itemy_k_prevedeni:
#9999     if line[0] == "Item":
#9999         continue
#9999     print(f'XXX {line[10]}')
#9999 
#9999     pdd = line[4].split(".")
#9999     pdd = datetime.date(int(pdd[-1]), int(pdd[1]), int(pdd[0]))
#9999     
#9999     if int(line[10]) < 0: # Pokud je Planned available na lince < 0.
#9999         print(line)
#9999         vrchol = line[0]
#9999        
#9999         #PZN105      
#9999         vrchol_available_qty_PZN105 = 0
#9999 
#9999         if database_pzn105_order_plan_dictionary.get(vrchol) != None:
#9999             print(database_pzn105_order_plan_dictionary.get(vrchol))
#9999 
#9999             for linka in range(1,len(database_pzn105_order_plan_dictionary.get(vrchol))+1):
#9999                 # print(range(1,len(database_pzn105_order_plan_dictionary.get(vrchol))+1))
#9999                 # print(linka)
#9999                 # print(database_pzn105_order_plan_dictionary.get(vrchol).get(linka))
#9999                 order_type = database_pzn105_order_plan_dictionary.get(vrchol).get(linka).get("Order type txt")
#9999                 if order_type.upper() == "PLANNED PURCHASE ORDER":
#9999                     order_type = ""
#9999                     continue
#9999                 datum = database_pzn105_order_plan_dictionary.get(vrchol).get(linka).get("Date")
#9999                 balance = database_pzn105_order_plan_dictionary.get(vrchol).get(linka).get("Transaction type txt")
#9999                 quantity = float(database_pzn105_order_plan_dictionary.get(vrchol).get(linka).get("Quantity"))
#9999                 # print(quantity, type(quantity))
#9999                 if balance == "+ (Planned Receipt)":
#9999                     vrchol_available_qty_PZN105 += quantity
#9999                 elif balance == "- (Planned Issue)": 
#9999                     vrchol_available_qty_PZN105 -= quantity
#9999                 else:
#9999                     print('POZOR - ERROR v +/- balance na u itemu {vrchol} na lince {linka}.')
#9999                 if vrchol_available_qty_PZN105 >= 0 and datum >= pdd:
#9999                     break
#9999             if vrchol_available_qty_PZN105 >= 0 and datum >= pdd:
#9999                 line.append(f'{datum}, PA: {vrchol_available_qty_PZN105}')
#9999             else:
#9999                 line.append(f'Neexistuje')              
#9999        
#9999         #PZN100       
#9999         vrchol_available_qty_PZN100 = 0
#9999 
#9999         if database_pzn100_order_plan_dictionary.get(vrchol) != None:      
#9999             for linka in range(1,len(database_pzn100_order_plan_dictionary.get(vrchol))+1):
#9999                 order_type = database_pzn100_order_plan_dictionary.get(vrchol).get(linka).get("Order type txt")
#9999                 if order_type.upper() == "PLANNED PURCHASE ORDER":
#9999                     order_type = ""
#9999                     continue
#9999                 datum = database_pzn100_order_plan_dictionary.get(vrchol).get(linka).get("Date")
#9999                 balance = database_pzn100_order_plan_dictionary.get(vrchol).get(linka).get("Transaction type txt")
#9999                 quantity = float(database_pzn100_order_plan_dictionary.get(vrchol).get(linka).get("Quantity"))
#9999                 # print(quantity, type(quantity))
#9999                 if balance == "+ (Planned Receipt)":
#9999                     vrchol_available_qty_PZN100 += quantity
#9999                 elif balance == "- (Planned Issue)": 
#9999                     vrchol_available_qty_PZN100 -= quantity
#9999                 else:
#9999                     print('POZOR - ERROR v +/- balance na u itemu {vrchol} na lince {linka}.')
#9999                 if vrchol_available_qty_PZN100 >= 0 and datum >= pdd:
#9999                     break
#9999             if vrchol_available_qty_PZN100 >= 0 and datum >= pdd:
#9999                 line.append(f'{datum}, PA: {vrchol_available_qty_PZN100}')
#9999             else:
#9999                 line.append(f'Neexistuje')
#9999 
#9999 
#9999 for line in itemy_k_prevedeni:
#9999     row_to_print = []
#9999     for pole in line:
#9999         
#9999         # print(f'Index {line.index(pole)}')
#9999         # print(f'Pole {pole}')
#9999         # print(f'Typ pole {type(pole)}')
#9999         
#9999         row_to_print.append(str(pole))
#9999     # print(row_to_print[9:13])
#9999     i = 0
#9999     for cislo in row_to_print[8:13]:
#9999         # print(str(cislo).replace(".",","))
#9999         row_to_print[8+i] = str(cislo).replace(".",",")
#9999         i+=1
#9999     print("|".join(row_to_print))
#9999                  
#9999 
#9999 
#9999 
#9999 
#9999 
#9999 
#9999                 #          cell_created.value.date() > dnes-timedelta(4):
#9999                 # print(row, {vrchol})
#9999                 # print(cell_created.value)
#9999                 # print(type(cell_created.value))
#9999 
#9999             # if str(cell_item.value).strip() == vrchol:
#9999                 # print(row)
#9999             # print(f'Cell item value {str(cell_item.value).strip()}')
#9999             # print(type(str(cell_item.value).strip()))
#9999             # print(f'Cell who processed {str(cell_who_processed.value).strip()}')
#9999             # print(type(str(cell_who_processed.value).strip()))
#9999                 
#9999                     #str(cell_who_processed.value).strip() == "None" and str(cell_item.value).strip() == vrchol and :
#9999                     # print(f'ANO')
#9999 
#9999 
#9999 
#9999 
#9999 
#9999 # for linka in unikatni_vrcholy_ks:
#9999 #     print(linka)
#9999 
#9999 
#9999 # Dotaz na stav (realny shartage na PZN105:)
#9999 # for vrchol, qty, available, buyer in unikatni_vrcholy_ks:
#9999 #     print(vrchol)
#9999 #     if vrchol == "Item":
#9999 #         continue
#9999 #     vrchol_available_qty_PZN100 = 0
#9999 #     if database_pzn100_order_plan_dictionary.get(vrchol) != None:
#9999 #         for linka in database_pzn100_order_plan_dictionary.get(vrchol):
#9999 #             # print(linka)
#9999 #             datum = database_pzn100_order_plan_dictionary.get(vrchol).get(linka).get("Date")
#9999 #             if datum <= proverit_do_datumu:
#9999 #                 balance = database_pzn100_order_plan_dictionary.get(vrchol).get(linka).get("Transaction type txt")
#9999 #                 quantity = float(database_pzn100_order_plan_dictionary.get(vrchol).get(linka).get("Quantity"))                
#9999 #                 print(quantity, type(quantity))
#9999 #                 if balance == "+ (Planned Receipt)":
#9999 #                     vrchol_available_qty_PZN100 += quantity
#9999 #                 elif balance == "- (Planned Issue)": 
#9999 #                     vrchol_available_qty_PZN100 -= quantity
#9999 #                 else:
#9999 #                     print('POZOR - ERROR v +/- balance na u itemu {vrchol} na lince linka.')
#9999 #     for item in unikatni_vrcholy_ks:
#9999 #         if item[0] == vrchol:
#9999 #             item[2] = vrchol_available_qty_PZN100
#9999 #     vrchol_available_qty_PZN100 = 0
#9999 # 
#9999 # for linka in unikatni_vrcholy_ks:
#9999 #     print(linka)
#9999 
#9999 
#9999 
#9999 
#9999 
#9999 
#9999 ## # Suma ks k prevedeni pro unikatni vrcholy
#9999 ## unikatni_vrcholy_ks = {line[0] for line in itemy_k_prevedeni} # Unikatni set vrcholu ze seznamu itemu k prevedeni.
#9999 ## unikatni_vrcholy_ks = list(unikatni_vrcholy_ks) # Prevedeni na list.
#9999 ## for item in unikatni_vrcholy_ks: # Doplneni ke kazdemu polozku sum mnozstvi inicializovanou na 0.
#9999 ##     unikatni_vrcholy_ks[unikatni_vrcholy_ks.index(item)] = [item, 0, 0, database_pzn100_order_plan_dictionary.get(item).get(1).get("Buyer")] if database_pzn100_order_plan_dictionary.get(item) != None else [item, 0, 0, 0]
#9999 ## # print(unikatni_vrcholy_ks)
#9999 ## 
#9999 ## # Projiti seznamu itemu k prevedeni a seceteni jejich mnozstvi a doplneni jako suma k itemu do seznamu unikatnich vrcholu.
#9999 ## for item in unikatni_vrcholy_ks:
#9999 ##     for line in itemy_k_prevedeni:
#9999 ##         if line[0] == item[0]:
#9999 ##             item[1] += int(line[3])
#9999 ## unikatni_vrcholy_ks.insert(0, ["Item", "Sum Qty shortage PZN105", "Available on PZN100 na prevod", "Nakupci"])
#9999 ## # print(unikatni_vrcholy_ks)
#9999 ## 
#9999 ## # Doplneni zahlavi.
#9999 ## itemy_k_prevedeni.insert(0, zahlavi)
#9999 ## print()
#9999 ## print(f'Seznam PZN 105 shortage linek z Master planu.')
#9999 ## for line in itemy_k_prevedeni:
#9999 ##     print(line)
#9999 ## print()
#9999 ## print(f'Suma mnozstvi shortage jednolivych dilu na dalsich {proverit_pristich_dni} dnu (od {dnes.strftime("%d/%m/%Y").replace("/", ".")} do {proverit_do_datumu.strftime("%d/%m/%Y").replace("/", ".")}).')
#9999 ## for line in unikatni_vrcholy_ks:
#9999 ##     print(line)
#9999 ## 